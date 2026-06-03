"""Tests for the community-spreadsheet sync (:mod:`arena.sync`).

The tests build a synthetic XLSX in memory with the exact layout the
parser expects (``Hot Tips`` sheet, fighter names in col A rows 2-8,
``Updated:`` date in cell M1, submitter names in the tip columns). Going
through real ``openpyxl`` round-trips means we exercise the parser exactly
as production will.

Coverage targets:

* parser:
    - reads ``Hot Tips`` sheet, ignores other sheets;
    - handles ``Updated: YYYY-MM-DD`` string and real date cells in M1;
    - classifies ``+5%`` / ``-5%`` headers as fighter tips;
    - classifies ``vs <Name>`` / ``+10% vs <Name>`` headers as matchup tips;
    - skips blank submitter cells.

* apply_sheet:
    - date-mismatch is a no-op skip, not an error;
    - additive merge respects an already-active manual tip
      (no submitter overwrite);
    - additive merge leaves a manually-active tip in place even when the
      sheet doesn't list it;
    - 15-tip cap enforced (excess sheet tips skipped, recorded);
    - idempotent: second run adds zero.

* sync_spreadsheet_once:
    - empty share URL -> ERROR status, no fetch attempted;
    - status / counts written into ``SpreadsheetSyncConfig``.

* API endpoint:
    - anonymous POST is rejected;
    - authenticated POST triggers a sync (we monkey-patch
      ``sync_spreadsheet_once`` so no network is touched).
"""
from __future__ import annotations

import datetime
from io import BytesIO

import pytest
import time_machine
from django.contrib.auth import get_user_model
from django.urls import reverse
from rest_framework.test import APIClient

from arena.models import (
    SPREADSHEET_SYNC_USERNAME,
    DAILY_TIP_CAP,
    DailyTipAuditLog,
    DailyTipSelection,
    SpreadsheetSyncConfig,
    TipDefinition,
)
from arena.sync import (
    ApplyResult,
    FetchError,
    ParsedSheet,
    SheetTip,
    apply_sheet,
    parse_hot_tips,
    sync_spreadsheet_once,
)


pytestmark = pytest.mark.django_db


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------


def _build_workbook(
    *,
    updated: str | datetime.date | None = "Updated: 2026-05-30",
    rows: list[dict] | None = None,
    sheet_name: str = "Hot Tips",
    extra_sheets: list[str] | None = None,
) -> bytes:
    """Produce an XLSX with the layout the parser expects.

    ``rows`` is a list of ``{"fighter": str, "tips": {header: submitter}}``
    dicts. The headers go into row 1 (cols B..); the fighter goes into
    col A; the submitter goes into the matching column.

    ``updated`` lands in cell M1 (string or date — both supported).
    """
    from openpyxl import Workbook

    wb = Workbook()
    # The default first sheet is renamed if needed.
    ws = wb.active
    ws.title = sheet_name

    if rows is None:
        rows = []

    # Collect headers in the order they first appear across rows so the
    # synthetic sheet stays compact.
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for h in row["tips"].keys():
            if h not in seen:
                headers.append(h)
                seen.add(h)

    # Row 1: A = "Fighter", then headers in B..
    ws.cell(row=1, column=1, value="Fighter")
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=1, column=idx, value=header)

    # Cell M1: updated date
    if updated is not None:
        ws["M1"] = updated

    # Rows 2..: fighter + submitter cells
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["fighter"])
        for header, submitter in row["tips"].items():
            col = 2 + headers.index(header)
            ws.cell(row=r, column=col, value=submitter)

    # Optional decoy sheets to prove the parser ignores them
    for name in extra_sheets or []:
        wb.create_sheet(name).cell(row=1, column=1, value="garbage")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def today():
    return datetime.date(2026, 5, 30)


# ---------------------------------------------------------------------------
# Parser tests
# ---------------------------------------------------------------------------


class TestParseHotTips:
    def test_string_updated_date_parses(self, seeded):
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[
                {"fighter": "Corrrak", "tips": {"+5%": "Night"}},
            ],
        )
        parsed = parse_hot_tips(xlsx)
        assert parsed.updated_date == datetime.date(2026, 5, 30)
        assert parsed.tips == (
            SheetTip(
                fighter_name="Corrrak",
                tip_type="fighter",
                modifier=5,
                opponent_name=None,
                submitter="Night",
            ),
        )

    def test_real_date_cell_parses(self, seeded):
        xlsx = _build_workbook(
            updated=datetime.date(2026, 5, 30),
            rows=[{"fighter": "Corrrak", "tips": {"-5%": "Aleks"}}],
        )
        parsed = parse_hot_tips(xlsx)
        assert parsed.updated_date == datetime.date(2026, 5, 30)
        assert parsed.tips[0].modifier == -5

    def test_matchup_header_with_vs(self, seeded):
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[
                {
                    "fighter": "Corrrak",
                    "tips": {"+10% vs Dura": "Night"},
                },
                # Dura must appear in col A so the classifier recognises
                # ``Dura`` as a fighter name for the matchup tip.
                {"fighter": "Dura", "tips": {}},
            ],
        )
        parsed = parse_hot_tips(xlsx)
        assert parsed.tips == (
            SheetTip(
                fighter_name="Corrrak",
                tip_type="matchup",
                modifier=10,
                opponent_name="Dura",
                submitter="Night",
            ),
        )

    def test_matchup_header_bare_vs_x(self, seeded):
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[
                {"fighter": "Corrrak", "tips": {"vs Dura": "Night"}},
                {"fighter": "Dura", "tips": {}},
            ],
        )
        parsed = parse_hot_tips(xlsx)
        # ``vs Dura`` alone has no explicit modifier — the parser
        # contractually maps any ``vs X`` header to the +10 matchup tip
        # per SPEC.md section 11.3 (matchup modifier is always +10).
        assert parsed.tips[0].tip_type == "matchup"
        assert parsed.tips[0].modifier == 10
        assert parsed.tips[0].opponent_name == "Dura"

    def test_real_layout_bare_fighter_name_columns(self, seeded):
        """The actual community sheet uses bare ``5`` / ``-5`` /
        ``<FighterName>`` headers, where a row × col cell means "row
        fighter has the tip versus the column fighter".
        """
        xlsx = _build_workbook(
            updated=datetime.date(2026, 5, 30),
            rows=[
                # Corrrak's row: +10% vs Leo, submitter Night
                {"fighter": "Corrrak", "tips": {"Leo": "Night"}},
                # Dura's row: +5% (any non-empty marker)
                {"fighter": "Dura", "tips": {"5": "X"}},
                # Leo's row: -5% submitter Aleks, AND +10% vs Dura by Night
                {"fighter": "Leo", "tips": {"-5": "Aleks", "Dura": "Night"}},
                # Otis's row: +10% vs both Leo and Ushug
                {"fighter": "Otis", "tips": {"Leo": "Night", "Ushug": "X"}},
                # Other fighters need to appear in col A so the classifier
                # recognises their names as matchup headers.
                {"fighter": "Gloz", "tips": {}},
                {"fighter": "Ushug", "tips": {}},
                {"fighter": "Viz", "tips": {}},
            ],
        )
        parsed = parse_hot_tips(xlsx)
        # Five tips: 1 matchup (Corrrak>Leo), 1 fighter (+5 Dura),
        # 1 fighter (-5 Leo), 1 matchup (Leo>Dura), 2 matchups (Otis>Leo
        # and Otis>Ushug). Total: 6.
        triples = sorted(
            (t.fighter_name, t.tip_type, t.modifier, t.opponent_name)
            for t in parsed.tips
        )
        assert triples == sorted(
            [
                ("Corrrak", "matchup", 10, "Leo"),
                ("Dura", "fighter", 5, None),
                ("Leo", "fighter", -5, None),
                ("Leo", "matchup", 10, "Dura"),
                ("Otis", "matchup", 10, "Leo"),
                ("Otis", "matchup", 10, "Ushug"),
            ]
        )

    def test_noise_columns_are_ignored(self, seeded):
        """Headers like ``Column1`` or ``Updated:`` outside the tip grid
        must classify as unknown and produce no tips.
        """
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[
                {
                    "fighter": "Corrrak",
                    "tips": {
                        "5": "Night",
                        "Column1": "Corrrak > Leo",
                        "Notes": "ignore me",
                    },
                },
            ],
        )
        parsed = parse_hot_tips(xlsx)
        assert len(parsed.tips) == 1
        assert parsed.tips[0].tip_type == "fighter"

    def test_blank_cells_are_skipped(self, seeded):
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[
                {
                    "fighter": "Corrrak",
                    "tips": {"+5%": "Night", "-5%": ""},
                },
                {
                    "fighter": "Dura",
                    "tips": {"+5%": None},
                },
            ],
        )
        parsed = parse_hot_tips(xlsx)
        # Only Corrrak +5% (Night) is non-blank.
        assert [t.fighter_name for t in parsed.tips] == ["Corrrak"]
        assert parsed.tips[0].modifier == 5

    def test_decoy_sheets_are_ignored(self, seeded):
        xlsx = _build_workbook(
            updated="Updated: 2026-05-30",
            rows=[{"fighter": "Corrrak", "tips": {"+5%": "Night"}}],
            extra_sheets=["Notes", "Calculations"],
        )
        parsed = parse_hot_tips(xlsx)
        assert len(parsed.tips) == 1

    def test_missing_hot_tips_sheet_raises(self, seeded):
        xlsx = _build_workbook(
            sheet_name="Wrong",
            updated="Updated: 2026-05-30",
            rows=[{"fighter": "Corrrak", "tips": {"+5%": "Night"}}],
        )
        from arena.sync import ParseError

        with pytest.raises(ParseError):
            parse_hot_tips(xlsx)


# ---------------------------------------------------------------------------
# apply_sheet tests
# ---------------------------------------------------------------------------


class TestApplySheet:
    def test_date_mismatch_is_skipped(self, seeded, today):
        parsed = ParsedSheet(
            updated_date=datetime.date(2026, 5, 1),
            tips=(
                SheetTip("Corrrak", "fighter", 5, None, "Night"),
            ),
        )
        result = apply_sheet(parsed, today)
        assert result.status == SpreadsheetSyncConfig.Status.SKIPPED
        assert result.added == 0
        assert DailyTipSelection.objects.count() == 0

    def test_no_updated_date_is_skipped(self, seeded, today):
        parsed = ParsedSheet(
            updated_date=None,
            tips=(SheetTip("Corrrak", "fighter", 5, None, "Night"),),
        )
        result = apply_sheet(parsed, today)
        assert result.status == SpreadsheetSyncConfig.Status.SKIPPED
        assert DailyTipSelection.objects.count() == 0

    def test_additive_merge_adds_missing_tips(self, seeded, today, tips):
        parsed = ParsedSheet(
            updated_date=today,
            tips=(
                SheetTip("Corrrak", "fighter", 5, None, "Night"),
                SheetTip("Dura", "fighter", -5, None, "Aleks"),
                SheetTip("Corrrak", "matchup", 10, "Dura", "Night"),
            ),
        )
        result = apply_sheet(parsed, today)
        assert result.status == SpreadsheetSyncConfig.Status.OK
        assert result.added == 3
        assert DailyTipSelection.objects.filter(date=today).count() == 3
        # All sync-added rows point at the system user.
        usernames = set(
            DailyTipSelection.objects.filter(date=today).values_list(
                "submitted_by__username", flat=True
            )
        )
        assert usernames == {SPREADSHEET_SYNC_USERNAME}
        externals = set(
            DailyTipSelection.objects.filter(date=today).values_list(
                "external_submitter_name", flat=True
            )
        )
        assert externals == {"Night", "Aleks"}
        # Audit log mirrors the activations.
        assert DailyTipAuditLog.objects.filter(date=today).count() == 3
        assert all(
            log.action == DailyTipAuditLog.Action.ACTIVATE
            for log in DailyTipAuditLog.objects.filter(date=today)
        )

    def test_manual_active_tip_is_not_touched(self, seeded, today, tips):
        """If a real user already activated a tip, the sync must NOT
        overwrite ``submitted_by`` or ``external_submitter_name``.
        """
        themys = get_user_model().objects.create_user(
            username="themys", password="x" * 16, first_name="Themys"
        )
        DailyTipSelection.objects.create(
            date=today,
            tip=tips["Corrrak +5%"],
            submitted_by=themys,
        )

        parsed = ParsedSheet(
            updated_date=today,
            tips=(SheetTip("Corrrak", "fighter", 5, None, "Night"),),
        )
        result = apply_sheet(parsed, today)
        assert result.status == SpreadsheetSyncConfig.Status.OK
        assert result.added == 0

        row = DailyTipSelection.objects.get(
            date=today, tip=tips["Corrrak +5%"]
        )
        assert row.submitted_by == themys
        assert row.external_submitter_name == ""

    def test_manual_tip_missing_from_sheet_stays_active(
        self, seeded, today, tips
    ):
        """Sync never deletes. A manually-active tip stays put even if the
        sheet doesn't mention it.
        """
        themys = get_user_model().objects.create_user(
            username="themys", password="x" * 16, first_name="Themys"
        )
        DailyTipSelection.objects.create(
            date=today,
            tip=tips["Corrrak +5%"],
            submitted_by=themys,
        )
        parsed = ParsedSheet(
            updated_date=today,
            tips=(SheetTip("Dura", "fighter", 5, None, "Aleks"),),
        )
        apply_sheet(parsed, today)
        assert DailyTipSelection.objects.filter(
            date=today, tip=tips["Corrrak +5%"]
        ).exists()
        # And the sheet's tip was added alongside it.
        assert DailyTipSelection.objects.filter(
            date=today, tip=tips["Dura +5%"]
        ).exists()
        assert DailyTipSelection.objects.filter(date=today).count() == 2

    def test_conflicting_tips_both_active(self, seeded, today, tips):
        """Manual +5 + sheet -5 on the same fighter -> both active."""
        themys = get_user_model().objects.create_user(
            username="themys", password="x" * 16, first_name="Themys"
        )
        DailyTipSelection.objects.create(
            date=today,
            tip=tips["Corrrak +5%"],
            submitted_by=themys,
        )
        parsed = ParsedSheet(
            updated_date=today,
            tips=(SheetTip("Corrrak", "fighter", -5, None, "Night"),),
        )
        result = apply_sheet(parsed, today)
        assert result.added == 1
        assert (
            DailyTipSelection.objects.filter(
                date=today,
                tip__in=[tips["Corrrak +5%"], tips["Corrrak -5%"]],
            ).count()
            == 2
        )

    def test_cap_blocks_excess_sheet_tips(self, seeded, today, tips):
        # Pre-fill with 14 manually-activated MATCHUP tips so headroom is
        # exactly 1 and our sheet's FIGHTER tips can't collide with the
        # pre-fill set.
        themys = get_user_model().objects.create_user(
            username="themys", password="x" * 16, first_name="Themys"
        )
        all_tips = list(tips.values())
        matchup_tip_defs = [
            t
            for t in all_tips
            if t.tip_type == TipDefinition.TipType.MATCHUP
        ]
        fighter_tip_defs = [
            t
            for t in all_tips
            if t.tip_type == TipDefinition.TipType.FIGHTER
        ]
        prefill = matchup_tip_defs[: DAILY_TIP_CAP - 1]
        assert len(prefill) == DAILY_TIP_CAP - 1
        for tip in prefill:
            DailyTipSelection.objects.create(
                date=today, tip=tip, submitted_by=themys
            )

        # Sheet wants to add 3 fighter tips. Only one should land; the
        # other two should be recorded in ``skipped_capped``.
        sheet_tip_defs = fighter_tip_defs[:3]
        assert len(sheet_tip_defs) == 3
        sheet_tips = tuple(
            SheetTip(
                fighter_name=td.fighter.name,
                tip_type="fighter",
                modifier=td.modifier,
                opponent_name=None,
                submitter="Night",
            )
            for td in sheet_tip_defs
        )

        parsed = ParsedSheet(updated_date=today, tips=sheet_tips)
        result = apply_sheet(parsed, today)
        assert result.status == SpreadsheetSyncConfig.Status.OK
        assert result.added == 1
        assert len(result.skipped_capped) == 2
        assert (
            DailyTipSelection.objects.filter(date=today).count()
            == DAILY_TIP_CAP
        )

    def test_unresolved_fighter_is_skipped(self, seeded, today):
        parsed = ParsedSheet(
            updated_date=today,
            tips=(
                SheetTip(
                    fighter_name="Nonexistent",
                    tip_type="fighter",
                    modifier=5,
                    opponent_name=None,
                    submitter="Night",
                ),
            ),
        )
        result = apply_sheet(parsed, today)
        assert result.added == 0
        assert len(result.skipped_unresolved) == 1
        assert DailyTipSelection.objects.count() == 0

    def test_idempotent(self, seeded, today, tips):
        parsed = ParsedSheet(
            updated_date=today,
            tips=(
                SheetTip("Corrrak", "fighter", 5, None, "Night"),
                SheetTip("Dura", "fighter", -5, None, "Aleks"),
            ),
        )
        first = apply_sheet(parsed, today)
        assert first.added == 2
        second = apply_sheet(parsed, today)
        assert second.added == 0
        assert DailyTipSelection.objects.filter(date=today).count() == 2


# ---------------------------------------------------------------------------
# sync_spreadsheet_once tests
# ---------------------------------------------------------------------------


class TestSyncSpreadsheetOnce:
    def test_empty_share_url_records_error(self, seeded):
        config = SpreadsheetSyncConfig.get_solo()
        config.share_url = ""
        config.save()

        result = sync_spreadsheet_once()
        assert result.status == SpreadsheetSyncConfig.Status.ERROR

        config.refresh_from_db()
        assert config.last_status == SpreadsheetSyncConfig.Status.ERROR
        assert "not configured" in config.last_message

    def test_disabled_records_skipped(self, seeded):
        config = SpreadsheetSyncConfig.get_solo()
        config.share_url = "https://example.invalid/share"
        config.enabled = False
        config.save()

        result = sync_spreadsheet_once()
        assert result.status == SpreadsheetSyncConfig.Status.SKIPPED

        config.refresh_from_db()
        assert "disabled" in config.last_message

    def test_force_bypasses_disabled(self, seeded, today, tips, monkeypatch):
        # force=True ignores the master switch but still uses the date gate.
        config = SpreadsheetSyncConfig.get_solo()
        config.share_url = "https://example.invalid/share"
        config.enabled = False
        config.save()

        # Stub the network: pretend OneDrive returned a valid sheet.
        xlsx = _build_workbook(
            updated=today,
            rows=[{"fighter": "Corrrak", "tips": {"+5%": "Night"}}],
        )
        monkeypatch.setattr(
            "arena.sync.fetch_workbook", lambda url: xlsx
        )

        # ``sync_spreadsheet_once`` calls ``current_game_day`` internally,
        # which reads the wall clock. Pin it to noon UTC on the same
        # day as the sheet (which is well after the NY midnight rollover
        # for both EDT and EST, so the date gate matches regardless of
        # season).
        with time_machine.travel(
            f"{today.isoformat()}T12:00:00+00:00", tick=False
        ):
            result = sync_spreadsheet_once(force=True)
        assert result.status == SpreadsheetSyncConfig.Status.OK
        assert result.added == 1

    def test_fetch_error_recorded(self, seeded, monkeypatch):
        config = SpreadsheetSyncConfig.get_solo()
        config.share_url = "https://example.invalid/share"
        config.enabled = True
        config.save()

        def boom(url):
            raise FetchError("simulated network failure")

        monkeypatch.setattr("arena.sync.fetch_workbook", boom)
        result = sync_spreadsheet_once()
        assert result.status == SpreadsheetSyncConfig.Status.ERROR
        assert "simulated network failure" in result.message


# ---------------------------------------------------------------------------
# API endpoint tests
# ---------------------------------------------------------------------------


class TestSyncEndpoint:
    def test_anonymous_post_rejected(self, seeded):
        response = APIClient().post(reverse("arena-sync"))
        assert response.status_code in (401, 403)

    def test_authenticated_post_triggers_sync(
        self, seeded, today, monkeypatch
    ):
        captured = {}

        def stub(force: bool = False):
            captured["force"] = force
            return ApplyResult(
                status=SpreadsheetSyncConfig.Status.OK,
                message="stub",
                sheet_date=today,
                added=0,
            )

        monkeypatch.setattr("arena.sync.sync_spreadsheet_once", stub)

        alice = get_user_model().objects.create_user(
            username="alice", password="x" * 16, first_name="Alice"
        )
        client = APIClient()
        client.force_authenticate(alice)
        response = client.post(reverse("arena-sync"))
        assert response.status_code == 200
        assert captured == {"force": True}
        # Response is the full arena state, including sync_status block.
        body = response.json()
        assert "sync_status" in body
        assert "active_tips" in body

    def test_state_endpoint_includes_sync_status(self, seeded):
        response = APIClient().get(reverse("arena-state"))
        assert response.status_code == 200
        body = response.json()
        assert "sync_status" in body
        ss = body["sync_status"]
        assert ss is not None
        assert "status" in ss
        assert "enabled" in ss
        assert "last_run_at" in ss
