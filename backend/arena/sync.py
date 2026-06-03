"""Community-spreadsheet sync.

The community maintains a shared OneDrive XLSX that lists today's hot tips
in a 7-fighter x N-column matrix on the ``Hot Tips`` sheet, with the
"as-of" date in cell ``M1``. This module fetches that file, parses it, and
*additively* merges the tips into our :class:`DailyTipSelection` table.

OneDrive personal "anyone with the link can view" shares are unfortunately
not accessible via the documented ``api.onedrive.com/shares/u!{b64}``
endpoint for files whose owner hasn't explicitly enabled public Graph
access (the endpoint returns 401 Unauthenticated). The browser-friendly
path is a two-step dance:

1. GET the share URL with a browser User-Agent and a cookie jar. The
   response is the HTML viewer for the file (sets a ``FedAuth`` anonymous
   session cookie).
2. The HTML embeds a signed direct-download URL — something like
   ``https://my.microsoftpersonalcontent.com/personal/<sitehex>/_layouts/
   15/download.aspx?UniqueId=<guid>&Translate=false&tempauth=v1e.<jwt>``
   inside a JSON-encoded JS payload (with ampersands escaped as
   ``\\u0026``). We extract it via regex, unescape, and GET it with the
   same cookie jar to get the raw XLSX bytes.

This is the same flow the browser uses anonymously; we just impersonate
it. No OAuth/Graph credentials required.

Design contracts (decided in the planning chat):

* **Date gating** — we only apply the sheet if the ``Updated:`` date in
  ``M1`` matches today's game day in ``America/New_York`` (see
  :func:`arena.services.current_game_day`). A stale sheet is treated as a
  no-op, not an error.

* **Manual wins** — the sync is strictly additive. It will:

    - add tips that are in the sheet but **not** currently active in our
      ``DailyTipSelection`` table;
    - leave already-active tips alone (no submitter overwrite);
    - **never** delete a tip, even if it's missing from the sheet.

  So if a real user manually clicked ``Corrrak +5%`` and the sheet says
  ``Corrrak -5%``, the post-sync state has *both* tips active. Conflicting
  tips are allowed by design — the calculator just sums them.

* **15-tip cap** — we respect the existing ``DAILY_TIP_CAP``. Tips from the
  sheet are accepted in the order the parser yields them (left-to-right,
  top-to-bottom) until the pool is full; the rest are recorded as
  ``last_skipped_count``.

* **Submitter attribution** — every sync-added row's ``submitted_by`` FK
  points at the ``spreadsheet-sync`` system user. The contributor's real
  name (read from the sheet cell) lives in
  ``DailyTipSelection.external_submitter_name`` so the UI can render it as
  ``"<name> (Spreadsheet)"`` without us creating a Django account per
  community contributor.

* **Idempotency** — calling :func:`sync_spreadsheet_once` twice in a row
  with the same sheet contents adds zero rows on the second call.

The OneDrive fetch goes through Microsoft's public ``shares/u!{b64}``
endpoint which accepts the share URL as a base64-url-encoded path and
returns the raw file bytes after a 302. That works for any
"anyone-with-the-link can view" share, which is how the sheet is
configured.
"""
from __future__ import annotations

import base64
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Iterable

import requests
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

from .models import (
    DAILY_TIP_CAP,
    SPREADSHEET_SYNC_USERNAME,
    DailyTipAuditLog,
    DailyTipSelection,
    Fighter,
    SpreadsheetSyncConfig,
    TipDefinition,
)
from .services import current_game_day


log = logging.getLogger(__name__)

# Hard limit on the XLSX payload so a hostile redirect can't make us
# allocate gigabytes. The real sheet is well under 50 KB.
MAX_XLSX_BYTES = 5 * 1024 * 1024  # 5 MiB
FETCH_TIMEOUT_SECONDS = 20
# Mimicking a browser User-Agent makes OneDrive serve the file directly
# instead of bouncing us through an HTML "Sign in" page on some accounts.
BROWSER_UA = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)

HOT_TIPS_SHEET_NAME = "Hot Tips"
# A2:J8 = 7 fighters x 10 columns (col A = fighter name, cols B-J = tip
# headers). We tolerate fewer columns; we hard-stop at row 8 / col J.
HOT_TIPS_FIRST_ROW = 2
HOT_TIPS_LAST_ROW = 8
HOT_TIPS_LAST_COL = 10  # column J
UPDATED_CELL = "M1"


class SyncError(Exception):
    """Top-level sync failure that should be surfaced to the operator."""


class FetchError(SyncError):
    """Network / OneDrive / HTTP problem reaching the spreadsheet."""


class ParseError(SyncError):
    """The XLSX downloaded fine but doesn't look like the expected sheet."""


@dataclass(frozen=True)
class SheetTip:
    """One tip cell, parsed but not yet resolved to a TipDefinition row."""

    fighter_name: str
    tip_type: str  # 'fighter' or 'matchup'
    modifier: int
    opponent_name: str | None
    submitter: str


@dataclass(frozen=True)
class ParsedSheet:
    updated_date: date | None
    tips: tuple[SheetTip, ...]


@dataclass
class ApplyResult:
    """Outcome of one sync run, written back into ``SpreadsheetSyncConfig``."""

    status: str  # SpreadsheetSyncConfig.Status value
    message: str = ""
    sheet_date: date | None = None
    added: int = 0
    skipped_capped: list[SheetTip] = field(default_factory=list)
    skipped_unresolved: list[tuple[SheetTip, str]] = field(default_factory=list)
    skipped_inactive_tip: list[SheetTip] = field(default_factory=list)

    @property
    def skipped_count(self) -> int:
        return (
            len(self.skipped_capped)
            + len(self.skipped_unresolved)
            + len(self.skipped_inactive_tip)
        )


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------

# XLSX files are ZIP archives -- every one starts with the PK\x03\x04 local
# file header. Used to distinguish "OneDrive served us the file" from
# "OneDrive served us an HTML sign-in page".
_XLSX_MAGIC = b"PK\x03\x04"

# The signed download URL we extract from the viewer HTML. Microsoft uses
# the ``my.microsoftpersonalcontent.com`` host for the actual file blob;
# the URL is JSON-embedded so ampersands appear as ``\u0026``.
_DOWNLOAD_URL_RE = re.compile(
    r"https://my\.microsoftpersonalcontent\.com/[^\s\"<>]+?download\.aspx\?[^\s\"<>]+",
    re.IGNORECASE,
)


def _encode_share_url(share_url: str) -> str:
    """Microsoft-style sharing URL encoding: ``u!`` + url-safe base64 (no =).

    Documented at
    https://learn.microsoft.com/en-us/onedrive/developer/rest-api/api/shares_get
    """
    b64 = base64.urlsafe_b64encode(share_url.encode("utf-8")).rstrip(b"=")
    return "u!" + b64.decode("ascii")


def _read_stream_into_bytes(response: requests.Response) -> bytes:
    """Buffer a response body into bytes, enforcing :data:`MAX_XLSX_BYTES`."""
    buf = bytearray()
    for chunk in response.iter_content(chunk_size=64 * 1024):
        if not chunk:
            continue
        buf.extend(chunk)
        if len(buf) > MAX_XLSX_BYTES:
            raise FetchError(
                f"sheet response exceeded {MAX_XLSX_BYTES} bytes; "
                "refusing to load"
            )
    return bytes(buf)


def _try_shares_endpoint(share_url: str, session: requests.Session) -> bytes | None:
    """Try the documented ``shares/u!{b64}/root/content`` endpoint.

    Returns the XLSX bytes on success, ``None`` on any non-200 outcome so
    the caller can fall back to the HTML-scrape path.
    """
    encoded = _encode_share_url(share_url)
    api_url = f"https://api.onedrive.com/v1.0/shares/{encoded}/root/content"
    try:
        response = session.get(
            api_url,
            headers={"User-Agent": BROWSER_UA},
            timeout=FETCH_TIMEOUT_SECONDS,
            allow_redirects=True,
            stream=True,
        )
    except requests.RequestException:
        return None
    with response:
        if response.status_code != 200:
            return None
        data = _read_stream_into_bytes(response)
        if data.startswith(_XLSX_MAGIC):
            return data
        return None


def _fetch_via_html_dance(
    share_url: str, session: requests.Session
) -> bytes:
    """Two-step anonymous fetch: open the viewer, extract the signed URL.

    Step 1 GETs the share URL with a browser UA. The response is the HTML
    viewer for the file (and sets a ``FedAuth`` cookie on the session).

    Step 2 unescapes ``\\u0026`` -> ``&`` in the HTML, regex-extracts a
    ``https://my.microsoftpersonalcontent.com/.../download.aspx?...&
    tempauth=...`` URL, and GETs that with the same session (cookie jar
    preserved).
    """
    try:
        viewer = session.get(
            share_url,
            headers={"User-Agent": BROWSER_UA},
            timeout=FETCH_TIMEOUT_SECONDS,
            allow_redirects=True,
        )
    except requests.RequestException as exc:
        raise FetchError(f"network error opening viewer: {exc}") from exc
    if viewer.status_code != 200:
        raise FetchError(
            f"viewer returned HTTP {viewer.status_code}; "
            "check the share URL is publicly viewable"
        )

    # The signed URL is embedded inside JSON inside HTML; ``&`` is escaped
    # as ``\u0026`` and ``/`` sometimes as ``\/``. Unescape both before
    # the regex hits.
    html = viewer.text.replace("\\u0026", "&").replace("\\/", "/")
    match = _DOWNLOAD_URL_RE.search(html)
    if match is None:
        raise FetchError(
            "could not find a signed download URL in the viewer HTML; "
            "OneDrive layout may have changed, or the share is not "
            "publicly viewable"
        )
    signed_url = match.group(0)

    try:
        download = session.get(
            signed_url,
            headers={
                "User-Agent": BROWSER_UA,
                "Referer": "https://onedrive.live.com/",
            },
            timeout=FETCH_TIMEOUT_SECONDS,
            allow_redirects=True,
            stream=True,
        )
    except requests.RequestException as exc:
        raise FetchError(
            f"network error downloading workbook: {exc}"
        ) from exc
    with download:
        if download.status_code != 200:
            raise FetchError(
                f"signed download returned HTTP {download.status_code}"
            )
        data = _read_stream_into_bytes(download)
    if not data.startswith(_XLSX_MAGIC):
        raise FetchError(
            "downloaded payload is not an XLSX file "
            f"(starts with {data[:8]!r})"
        )
    return data


def fetch_workbook(share_url: str) -> bytes:
    """Fetch the raw XLSX bytes for an "anyone-with-the-link" OneDrive share.

    Tries the public ``api.onedrive.com/shares/u!{b64}/root/content``
    endpoint first (single request, no HTML parsing). Falls back to the
    browser-style HTML viewer dance if the API endpoint says 401/404/etc
    — that's the common case for personal shares that haven't opted into
    Graph access.

    Raises :class:`FetchError` for any unrecoverable failure (network,
    HTTP, unparseable HTML, oversized response). The caller writes the
    error into :class:`SpreadsheetSyncConfig`.
    """
    if not share_url:
        raise FetchError("share_url is empty; nothing to fetch")

    session = requests.Session()
    fast = _try_shares_endpoint(share_url, session)
    if fast is not None:
        return fast
    return _fetch_via_html_dance(share_url, session)


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------


# Headers we recognise in row 1. The actual community sheet uses bare
# headers like ``5`` / ``-5`` / ``<FighterName>``; an earlier draft also
# supported ``+5%`` / ``-5%`` / ``vs <Name>`` / ``+10% vs <Name>`` so we
# keep accepting those too in case the sheet schema drifts.
_FIGHTER_PLUS_RE = re.compile(r"^\s*\+?\s*5\s*%?\s*$", re.IGNORECASE)
_FIGHTER_MINUS_RE = re.compile(r"^\s*[-\u2212]\s*5\s*%?\s*$", re.IGNORECASE)
_MATCHUP_VS_RE = re.compile(
    r"vs\.?\s*([A-Za-z][A-Za-z\-' ]*)", re.IGNORECASE
)


def _classify_header(
    header: str, fighter_names: frozenset[str]
) -> tuple[str, int, str | None] | None:
    """Decode a column header into (tip_type, modifier, opponent_name).

    ``fighter_names`` is the set of fighter names found in column A of the
    sheet — used to recognise the bare-name "fighter-vs-fighter" matchup
    columns (where the column header is just ``"Leo"`` and the row's
    fighter is the winner). Case-insensitive.

    Returns ``None`` for unknown / blank headers so the caller can skip
    the column entirely (also catches noise headers like ``"Column1"``,
    ``"Updated:"`` etc. that sit outside the tip grid).
    """
    text = (header or "").strip()
    if not text:
        return None

    # Bare fighter name = matchup column. Row fighter is the winner
    # (target_fighter = +10), column header is the loser (opponent).
    matched_name = _match_fighter_name(text, fighter_names)
    if matched_name is not None:
        return ("matchup", 10, matched_name)

    # ``vs <Name>`` / ``+10% vs <Name>`` (legacy format)
    vs_match = _MATCHUP_VS_RE.search(text)
    if vs_match:
        opponent = vs_match.group(1).strip()
        matched_name = _match_fighter_name(opponent, fighter_names)
        return ("matchup", 10, matched_name or opponent)

    if _FIGHTER_PLUS_RE.match(text):
        return ("fighter", 5, None)
    if _FIGHTER_MINUS_RE.match(text):
        return ("fighter", -5, None)
    return None


def _match_fighter_name(
    candidate: str, fighter_names: frozenset[str]
) -> str | None:
    """Case-insensitive lookup of ``candidate`` in ``fighter_names``.

    Returns the canonical (sheet-spelt) name on a match so downstream
    resolution against ``Fighter.name`` stays exact.
    """
    needle = candidate.casefold().strip()
    for name in fighter_names:
        if name.casefold() == needle:
            return name
    return None


def _parse_updated_cell(value) -> date | None:
    """Decode cell ``M1``'s contents into a date.

    The cell may hold a real date/datetime, or a free-form string like
    ``Updated: 2026-05-30`` / ``5/30/2026``. We strip an optional
    ``Updated:`` prefix and try a few common formats.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    s = re.sub(r"^\s*updated\s*:\s*", "", s, flags=re.IGNORECASE)
    s = s.strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%B %d, %Y",
        "%b %d, %Y",
        "%d %B %Y",
        "%d %b %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_hot_tips(xlsx_bytes: bytes) -> ParsedSheet:
    """Decode the ``Hot Tips`` sheet into a list of :class:`SheetTip`.

    Layout assumed:

    * Row 1: column headers (col A is "Fighter" / blank, cols B-J describe
      the tip in that column).
    * Rows 2..8: one fighter per row (col A is the fighter name; cols B-J
      hold the submitter's name when that fighter has the tip in that
      column, blank otherwise).
    * Cell M1: ``Updated: YYYY-MM-DD`` (or a real date cell).

    Other sheets in the workbook are ignored.
    """
    # ``openpyxl`` is heavy to import; defer until we actually need it
    # (keeps Django startup snappy and avoids importing it in management
    # commands that don't sync).
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dep missing in dev
        raise ParseError(
            "openpyxl is required to parse the spreadsheet; "
            "did you `uv sync`?"
        ) from exc

    try:
        wb = load_workbook(
            filename=io.BytesIO(xlsx_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise ParseError(f"could not open workbook: {exc}") from exc

    if HOT_TIPS_SHEET_NAME not in wb.sheetnames:
        raise ParseError(
            f"workbook has no {HOT_TIPS_SHEET_NAME!r} sheet "
            f"(found {wb.sheetnames!r})"
        )
    ws = wb[HOT_TIPS_SHEET_NAME]

    updated_date = _parse_updated_cell(ws[UPDATED_CELL].value)

    # Pass 1: discover the set of fighter names from column A so the
    # header classifier can recognise the bare-name matchup columns
    # (cols D..J in the live sheet are headed by other fighters' names).
    row_fighters: list[tuple[int, str]] = []
    for row in range(HOT_TIPS_FIRST_ROW, HOT_TIPS_LAST_ROW + 1):
        cell = ws.cell(row=row, column=1).value
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            row_fighters.append((row, name))
    fighter_names = frozenset(name for _row, name in row_fighters)

    # Pass 2: classify row-1 headers for columns B..HOT_TIPS_LAST_COL.
    headers: list[tuple[int, tuple[str, int, str | None]]] = []
    for col in range(2, HOT_TIPS_LAST_COL + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        classified = _classify_header(str(raw), fighter_names)
        if classified is None:
            # Noise column (e.g. ``Column1``, ``Updated:``, ``Notes``) —
            # silently skip; the operator's free to add scratch columns.
            continue
        headers.append((col, classified))

    # Pass 3: walk every fighter row and emit one SheetTip per non-empty
    # cell. The cell *value* is the submitter name (or a placeholder like
    # ``"X"`` meaning "active, no contributor recorded"). Diagonal
    # cells (fighter's own column) should naturally be blank and so get
    # skipped; we don't enforce it.
    tips: list[SheetTip] = []
    for row, fighter_name in row_fighters:
        for col, (tip_type, modifier, opponent) in headers:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue
            submitter = str(cell_value).strip()
            if not submitter:
                continue
            tips.append(
                SheetTip(
                    fighter_name=fighter_name,
                    tip_type=tip_type,
                    modifier=modifier,
                    opponent_name=opponent,
                    submitter=submitter,
                )
            )

    wb.close()
    return ParsedSheet(updated_date=updated_date, tips=tuple(tips))


# ---------------------------------------------------------------------------
# Apply (merge into DailyTipSelection)
# ---------------------------------------------------------------------------


def _get_system_user():
    """Return the ``spreadsheet-sync`` system user (created in migration 0002)."""
    User = get_user_model()
    try:
        return User.objects.get(username=SPREADSHEET_SYNC_USERNAME)
    except User.DoesNotExist as exc:  # pragma: no cover - migration enforces
        raise SyncError(
            f"system user {SPREADSHEET_SYNC_USERNAME!r} missing; "
            "did migrations run?"
        ) from exc


def _resolve_tip_definition(
    sheet_tip: SheetTip,
    fighters_by_name: dict[str, Fighter],
) -> TipDefinition | None:
    """Resolve a parsed :class:`SheetTip` to a real :class:`TipDefinition`.

    Returns ``None`` if no matching definition exists (e.g. a typo in the
    sheet, or a fighter we haven't seeded). The caller logs and skips.
    """
    from django.db.models import Q

    from .models import Matchup

    fighter = fighters_by_name.get(sheet_tip.fighter_name)
    if fighter is None:
        return None

    if sheet_tip.tip_type == "fighter":
        return (
            TipDefinition.objects.filter(
                tip_type=TipDefinition.TipType.FIGHTER,
                fighter=fighter,
                target_fighter=fighter,
                modifier=sheet_tip.modifier,
                is_active=True,
            )
            .first()
        )

    # Matchup tip — find the (fighter, opponent) matchup row in either
    # orientation, then the TipDefinition that targets ``fighter`` in it.
    opponent = fighters_by_name.get(sheet_tip.opponent_name or "")
    if opponent is None:
        return None
    matchup_pks = Matchup.objects.filter(
        (Q(fighter_a=fighter) & Q(fighter_b=opponent))
        | (Q(fighter_a=opponent) & Q(fighter_b=fighter))
    ).values_list("pk", flat=True)
    return (
        TipDefinition.objects.filter(
            tip_type=TipDefinition.TipType.MATCHUP,
            target_fighter=fighter,
            modifier=sheet_tip.modifier,
            is_active=True,
            matchup__in=matchup_pks,
        )
        .first()
    )


def apply_sheet(parsed: ParsedSheet, game_day: date) -> ApplyResult:
    """Additively merge a parsed sheet into ``DailyTipSelection`` for ``game_day``.

    See module docstring for the merge contract. Returns a populated
    :class:`ApplyResult` describing exactly what happened so the API and
    admin can surface useful diagnostics.
    """
    # Date gate: ignore stale sheets.
    if parsed.updated_date is None:
        return ApplyResult(
            status=SpreadsheetSyncConfig.Status.SKIPPED,
            message="sheet has no parseable Updated: date in M1",
            sheet_date=None,
        )
    if parsed.updated_date != game_day:
        return ApplyResult(
            status=SpreadsheetSyncConfig.Status.SKIPPED,
            message=(
                f"sheet date {parsed.updated_date.isoformat()} does not "
                f"match game day {game_day.isoformat()}"
            ),
            sheet_date=parsed.updated_date,
        )

    system_user = _get_system_user()
    fighters_by_name = {f.name: f for f in Fighter.objects.all()}

    result = ApplyResult(
        status=SpreadsheetSyncConfig.Status.OK,
        sheet_date=parsed.updated_date,
    )

    with transaction.atomic():
        # Snapshot today's selections under a single query so we have a
        # consistent view and (on Postgres) we lock the rows; same pattern
        # used by :class:`arena.views.ToggleTipView`.
        from django.db import connection

        base_qs = DailyTipSelection.objects.filter(date=game_day)
        if connection.features.has_select_for_update:
            base_qs = base_qs.select_for_update()
        existing = list(base_qs)
        existing_tip_ids = {s.tip_id for s in existing}
        current_count = len(existing)

        for sheet_tip in parsed.tips:
            tip_def = _resolve_tip_definition(sheet_tip, fighters_by_name)
            if tip_def is None:
                result.skipped_unresolved.append(
                    (
                        sheet_tip,
                        "no matching active TipDefinition "
                        "(fighter/opponent typo?)",
                    )
                )
                continue
            if tip_def.id in existing_tip_ids:
                # Already active in our DB — manual-wins says hands off.
                continue
            if current_count >= DAILY_TIP_CAP:
                result.skipped_capped.append(sheet_tip)
                continue
            DailyTipSelection.objects.create(
                date=game_day,
                tip=tip_def,
                submitted_by=system_user,
                external_submitter_name=sheet_tip.submitter[:150],
            )
            DailyTipAuditLog.objects.create(
                date=game_day,
                tip=tip_def,
                action=DailyTipAuditLog.Action.ACTIVATE,
                actor=system_user,
                external_actor_name=sheet_tip.submitter[:150],
            )
            existing_tip_ids.add(tip_def.id)
            current_count += 1
            result.added += 1

    parts: list[str] = [f"added {result.added}"]
    if result.skipped_capped:
        parts.append(f"skipped {len(result.skipped_capped)} (cap)")
    if result.skipped_unresolved:
        parts.append(f"skipped {len(result.skipped_unresolved)} (unresolved)")
    result.message = ", ".join(parts)
    return result


# ---------------------------------------------------------------------------
# Top-level orchestration
# ---------------------------------------------------------------------------


def sync_spreadsheet_once(*, force: bool = False) -> ApplyResult:
    """Run one fetch -> parse -> apply cycle and persist the outcome.

    ``force`` is reserved for the manual button: it bypasses the
    ``enabled`` flag check (so an admin can troubleshoot without flipping
    the master switch) but never bypasses the date gate.
    """
    config = SpreadsheetSyncConfig.get_solo()

    if not force and not config.enabled:
        result = ApplyResult(
            status=SpreadsheetSyncConfig.Status.SKIPPED,
            message="sync disabled by config",
        )
        _record_result(config, result)
        return result

    if not config.share_url:
        result = ApplyResult(
            status=SpreadsheetSyncConfig.Status.ERROR,
            message="share_url is not configured",
        )
        _record_result(config, result)
        return result

    try:
        xlsx_bytes = fetch_workbook(config.share_url)
        parsed = parse_hot_tips(xlsx_bytes)
    except SyncError as exc:
        result = ApplyResult(
            status=SpreadsheetSyncConfig.Status.ERROR,
            message=str(exc),
        )
        log.warning("spreadsheet sync failed: %s", exc)
        _record_result(config, result)
        return result
    except Exception as exc:  # pragma: no cover - belt-and-braces
        result = ApplyResult(
            status=SpreadsheetSyncConfig.Status.ERROR,
            message=f"unexpected error: {exc.__class__.__name__}: {exc}",
        )
        log.exception("spreadsheet sync crashed")
        _record_result(config, result)
        return result

    game_day = current_game_day()
    result = apply_sheet(parsed, game_day)
    _record_result(config, result)
    return result


def _record_result(
    config: SpreadsheetSyncConfig, result: ApplyResult
) -> None:
    config.last_run_at = timezone.now()
    config.last_status = result.status
    config.last_message = result.message[:5000]
    config.last_sheet_date = result.sheet_date
    config.last_added_count = result.added
    config.last_skipped_count = result.skipped_count
    config.save(
        update_fields=[
            "last_run_at",
            "last_status",
            "last_message",
            "last_sheet_date",
            "last_added_count",
            "last_skipped_count",
        ]
    )
